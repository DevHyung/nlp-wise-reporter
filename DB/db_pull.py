import pymysql.cursors
import time
from openpyxl import load_workbook, Workbook
import os
import CONFIG as config

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME): # 덮어씌우기
        if _DATA == None: return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for data in _DATA:
            sheet.append(data)
        book.save(_FILENAME)
    else:  # 새로 만듦
        if _HEADER == None: return None
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 5
        sheet.column_dimensions['G'].width = 10
        book.save(_FILENAME)


def show_time():
    now = time.localtime()
    print("%04d/%02d/%02d %02d:%02d:%02d" %
          (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec))


def log(_option, _msg):
    if   _option.lower() == 'e': print("[Error]   : {}".format(_msg))
    elif _option.lower() == 's': print("[Success] : {}".format(_msg))
    elif _option.lower() == 'i': print("[Info]    : {}".format(_msg))


def db_connect():
    # Connect to the database
    connection = pymysql.connect(host=config.host,
                                 port=config.port,
                                 user=config.user,
                                 password=config.password,
                                 db=config.db,
                                 charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
    return connection


def query(_sql):
    dataList = []
    try:
        with connection.cursor() as cursor:
            cursor.execute(_sql)
            results = cursor.fetchall()
            for r in results:
                dataList.append([r['_firstdate'], r['_finaldate'], r['_naverUrl'], r['_originalUrl'], r['_newsTitle'],
                                 r['_cate'], r['_crawldate'], r['_originalcontent'], r['_extcontent']])
    except Exception as e:
        connection.close()
        log('e', e)

    save_excel(FILENAME, dataList, None)
if __name__ == "__main__":
    FILENAME = "mostviewed.xlsx"
    headerList = ['_firstdate', '_finaldate', '_naverUrl', '_originalUrl', '_newsTitle',
                  '_cate', '_crawldate', '_originalcontent', '_extcontent']
    save_excel(FILENAME, None, headerList)

    # SQL
    sql = "SELECT * FROM mostviewed_newspaper;"

    # Connect.
    connection = db_connect()

    # Query.
    # 16857 줄일때 2020/02/25 17:20:04 - 2020/02/25 17:24:05
    show_time()
    query(sql)
    show_time()

    # Finally
    connection.close()